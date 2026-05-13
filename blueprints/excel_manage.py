import openpyxl
from flask import Blueprint, jsonify, request
from flask_login import current_user, login_required


def create_excel_manage_blueprint(
    db,
    get_json_data,
    require_fields,
    normalize_ip,
    normalize_mac,
    normalize_mode,
    normalize_vlan,
    get_switch_runtime_data,
    get_manager,
    assert_interface_not_protected,
    save_binding_state,
    format_switch_log,
    send_xlsx_workbook,
    autosize_worksheet,
    json_error,
    internal_error,
    permission_required,
):
    bp = Blueprint('excel_manage', __name__)

    @bp.route('/api/excel_binding_template', methods=['GET'])
    @login_required
    def download_excel_binding_template():
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = 'binding_template'
            headers = ['交换机IP', '端口', 'VLAN', '绑定IP', '绑定MAC', '模式']
            sheet.append(headers)
            sheet.append(['10.139.100.205', 'GE1/0/18', 202, '10.139.27.15', 'aaaa-bbbb-cccc', 'access'])
            sheet.append(['10.139.100.205', 'GE1/0/18', 202, '10.139.27.16', 'bbbb-cccc-dddd', 'access'])
            sheet.append(['10.139.100.213', 'XGE1/0/1', 199, '10.139.99.10', 'cccc-dddd-eeee', 'trunk'])
            sheet.freeze_panes = 'A2'
            colors = ['206BC4', '2FB344', 'F59F00', '17A2B8', 'D63939', '6F42C1']
            for idx, cell in enumerate(sheet[1]):
                cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
                cell.fill = openpyxl.styles.PatternFill('solid', fgColor=colors[idx])
            autosize_worksheet(sheet)
            return send_xlsx_workbook(wb, 'terminal_binding_batch_template.xlsx')
        except Exception as e:
            return internal_error('生成 Excel 批量部署模板失败，请稍后重试', e)

    @bp.route('/api/parse_excel', methods=['POST'])
    @login_required
    def parse_excel():
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'msg': '未找到上传的文件'})

        file = request.files['file']
        if file.filename == '':
            return jsonify({'status': 'error', 'msg': '文件名不能为空'})

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            sheet = wb.active

            headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
            required_cols = ['交换机IP', '端口', 'VLAN', '绑定IP', '绑定MAC', '模式']

            col_indices = {}
            for req in required_cols:
                if req in headers:
                    col_indices[req] = headers.index(req)
                else:
                    return jsonify({'status': 'error', 'msg': f"Excel 缺少必填列头：{req}"})

            data = []
            seen_keys = set()
            duplicate_count = 0
            switch_ips = set()
            port_keys = set()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                switch_ip = row[col_indices['交换机IP']]
                if not switch_ip:
                    continue
                switch_ip = normalize_ip(switch_ip, '交换机 IP')
                bind_ip = normalize_ip(row[col_indices['绑定IP']], '绑定 IP')
                mac = normalize_mac(row[col_indices['绑定MAC']])
                mode = normalize_mode(row[col_indices['模式']], '模式列')

                interface = str(row[col_indices['端口']]).strip()
                key = (switch_ip, interface, bind_ip, mac)
                if key in seen_keys:
                    duplicate_count += 1
                seen_keys.add(key)
                switch_ips.add(switch_ip)
                port_keys.add((switch_ip, interface))
                data.append({
                    'switch_ip': switch_ip,
                    'interface': interface,
                    'vlan': normalize_vlan(row[col_indices['VLAN']], 'VLAN 列'),
                    'bind_ip': bind_ip,
                    'mac': mac,
                    'mode': mode,
                })

            summary = {
                'rows': len(data),
                'switch_count': len(switch_ips),
                'port_task_count': len(port_keys),
                'duplicate_count': duplicate_count,
                'access_count': sum(1 for item in data if item['mode'] == 'access'),
                'trunk_count': sum(1 for item in data if item['mode'] == 'trunk'),
            }
            return jsonify({'status': 'success', 'data': data, 'summary': summary})
        except Exception as e:
            return internal_error('解析 Excel 失败，请检查文件格式和内容', e)

    @bp.route('/api/execute_excel_group', methods=['POST'])
    @login_required
    @permission_required('access.write')
    def execute_excel_group():
        try:
            data = get_json_data()
            rows = data.get('rows')
            if not isinstance(rows, list) or not rows:
                raise ValueError('请求中必须包含 rows 数组，且不能为空')

            client_ip = request.remote_addr
            normalized_rows = []
            switch_ips = set()
            interfaces = set()
            modes = set()

            for row in rows:
                if not isinstance(row, dict):
                    raise ValueError('rows 中存在无效记录')
                require_fields(row, ['switch_ip', 'interface', 'vlan', 'bind_ip', 'mac', 'mode'])
                item = {
                    'switch_ip': normalize_ip(row['switch_ip'], '交换机 IP'),
                    'interface': str(row['interface']).strip(),
                    'vlan': normalize_vlan(row['vlan']),
                    'bind_ip': normalize_ip(row['bind_ip'], '绑定 IP'),
                    'mac': normalize_mac(row['mac']),
                    'mode': normalize_mode(row['mode']),
                }
                if not item['interface']:
                    raise ValueError('端口不能为空')
                normalized_rows.append(item)
                switch_ips.add(item['switch_ip'])
                interfaces.add(item['interface'])
                modes.add(item['mode'])

            if len(switch_ips) != 1 or len(interfaces) != 1:
                raise ValueError('同一次批量下发只允许处理同一交换机的同一个端口')
            if len(modes) != 1:
                raise ValueError('同一端口的批量下发必须使用同一种模式')

            switch_ip = normalized_rows[0]['switch_ip']
            interface = normalized_rows[0]['interface']
            mode = normalized_rows[0]['mode']
            runtime = get_switch_runtime_data(switch_ip)
            mgr = get_manager(runtime)
            assert_interface_not_protected(mgr, interface)

            raw_log = mgr.configure_port_bindings_batch(interface, normalized_rows, mode)
            for row in normalized_rows:
                save_binding_state(
                    row['switch_ip'],
                    row['interface'],
                    row['vlan'],
                    row['bind_ip'],
                    row['mac'],
                    row['mode'],
                )

            vlan_summary = ','.join(sorted({row['vlan'] for row in normalized_rows}))
            details = f"[Excel批量聚合] 端口:{interface} | 条数:{len(normalized_rows)} | 模式:{mode} | VLAN:{vlan_summary}"
            db.log_operation(current_user.username, client_ip, switch_ip, "批量端口绑定", details, "成功")
            return jsonify({'status': 'success', 'log': format_switch_log(raw_log)})
        except ValueError as e:
            if 'client_ip' in locals():
                db.log_operation(
                    current_user.username,
                    client_ip,
                    normalized_rows[0]['switch_ip'] if 'normalized_rows' in locals() and normalized_rows else 'Unknown',
                    "批量端口绑定",
                    f"[Excel批量聚合] 报错: {str(e)}",
                    "失败",
                )
            return json_error(str(e))
        except Exception as e:
            if 'client_ip' in locals():
                db.log_operation(
                    current_user.username,
                    client_ip,
                    normalized_rows[0]['switch_ip'] if 'normalized_rows' in locals() and normalized_rows else 'Unknown',
                    "批量端口绑定",
                    f"[Excel批量聚合] 报错: {str(e)}",
                    "失败",
                )
            return internal_error('批量下发失败，请检查设备状态和表格内容', e)

    @bp.route('/api/execute_excel_row', methods=['POST'])
    @login_required
    @permission_required('access.write')
    def execute_excel_row():
        try:
            d = get_json_data()
            client_ip = request.remote_addr
            require_fields(d, ['switch_ip', 'interface', 'vlan', 'bind_ip', 'mac', 'mode'])
            switch_ip = normalize_ip(d.get('switch_ip'), '交换机 IP')
            interface = str(d.get('interface')).strip()
            vlan = normalize_vlan(d.get('vlan'))
            bind_ip = normalize_ip(d.get('bind_ip'), '绑定 IP')
            mac = normalize_mac(d.get('mac'))
            mode = normalize_mode(d.get('mode', 'access'))
            runtime = get_switch_runtime_data(switch_ip)
            mgr = get_manager(runtime)
            assert_interface_not_protected(mgr, interface)

            raw_log = mgr.configure_port_binding(interface, vlan, bind_ip, mac, mode)
            log_output = format_switch_log(raw_log)
            save_binding_state(switch_ip, interface, vlan, bind_ip, mac, mode)
            details = f"[Excel批量] 端口:{interface} | IP:{bind_ip} | MAC:{mac} | 模式:{mode} | VLAN:{vlan}"
            db.log_operation(current_user.username, client_ip, switch_ip, "批量端口绑定", details, "成功")

            return jsonify({'status': 'success', 'log': log_output})
        except ValueError as e:
            details = f"[Excel批量] 端口:{locals().get('interface', '')} | IP:{locals().get('bind_ip', '')} | MAC:{locals().get('mac', '')}"
            db.log_operation(current_user.username, client_ip, locals().get('switch_ip', 'Unknown'), "批量端口绑定", f"{details} | 报错: {str(e)}", "失败")
            return json_error(str(e))
        except Exception as e:
            details = f"[Excel批量] 端口:{interface} | IP:{bind_ip} | MAC:{mac}"
            db.log_operation(current_user.username, client_ip, switch_ip, "批量端口绑定", f"{details} | 报错: {str(e)}", "失败")
            return internal_error('批量下发失败，请检查设备状态和参数', e)

    return bp
