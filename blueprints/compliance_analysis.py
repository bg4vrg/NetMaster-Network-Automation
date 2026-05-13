import csv
import datetime
import ipaddress
import io
import re
from collections import defaultdict

import openpyxl
from flask import Blueprint, jsonify, request, send_file
from flask_login import current_user, login_required

import database as db
from oui_service import lookup_mac_vendor


def create_compliance_analysis_blueprint(internal_error):
    bp = Blueprint('compliance_analysis', __name__)

    TEMPLATE_COLUMNS = {
        'agent': ['IP地址', 'MAC地址', '主机名', '用户', '部门', '状态', '最近上线时间', '备注'],
        'registry': ['rule', '设备ip', '设备mac', '使用/责任人', '安装位置', '设备用途', '设备类型', '操作系统', '一机两用', '苏警盾安装', '惠安资产', '备注'],
    }

    TEMPLATE_SAMPLES = {
        'agent': ['10.139.1.20', '00e0-fc12-3456', 'PC-001', '张三', '信息科', '已保护', '2026-05-11 08:30:00', '9200平台导出可保留原字段'],
        'registry': [1, '10.139.1.20', '00:e0:fc:12:34:56', '张三', '1026室', '办公', '台式机', 'Windows 10', '已安装', '已安装', '已登记', '人工批准入网台账'],
    }

    FIELD_ALIASES = {
        'ip': ['ip', 'ip地址', '终端ip', '地址', 'ip address', '设备ip', '绑定时对应ip', '客户端通讯ip'],
        'mac': ['mac', 'mac地址', '物理地址', 'mac address', '设备mac', '查询mac地址'],
        'last_seen': ['最近一次上线时间', '最后上线时间', '更新时间', '最后在线时间', 'last_seen', 'last seen'],
        'switch_name': ['接入设备', '交换机', '接入交换机', '交换机名称', '设备名称'],
        'switch_ip': ['交换机地址', '交换机ip', '接入交换机ip'],
        'port': ['端口', '接入端口', '接口'],
        'vlan': ['vlan', 'VLAN'],
        'status': ['在线状态', '状态'],
        'group': ['分组', '网络', '区域'],
        'owner': ['用户', '使用人', '责任人', '使用/责任人', '姓名'],
        'department': ['部门', '单位', '科室', '科室名称', '组织机构'],
        'device_type': ['设备类型', '设备类型(参考)', '资产类型', '类型', '计算机类型'],
        'install_location': ['安装位置', '计算机所在地', '房间号', '位置'],
        'device_purpose': ['设备用途', '用途'],
        'asset_status': ['惠安资产', '资产状态', '资产登记'],
        'note': ['备注', '设备描述', '说明'],
        'core_version': ['客户端核心版本号', '核心版本号', '客户端版本', '版本号'],
        'install_status': ['安装状态', '是否安装', '客户端安装状态'],
        'protection_status': ['保护状态', '一机两用', '苏警盾安装', '信任状态'],
        'hostname': ['主机名', '计算机名', '终端名称', '设备名称'],
    }

    def normalize_header(value):
        return re.sub(r'\s+', '', str(value or '').strip()).lower()

    def guess_field(headers, key):
        normalized = {normalize_header(header): header for header in headers}
        for alias in FIELD_ALIASES.get(key, []):
            hit = normalized.get(normalize_header(alias))
            if hit:
                return hit
        return ''

    def normalize_ip(value):
        raw = str(value or '').strip()
        if not raw:
            return ''
        try:
            ip = ipaddress.ip_address(raw)
            if ip.version != 4:
                return ''
            text = str(ip)
            if text == '0.0.0.0' or text.startswith('169.254.'):
                return ''
            return text
        except Exception:
            return ''

    def normalize_mac(value):
        raw = str(value or '').strip()
        if not raw:
            return ''
        compact = re.sub(r'[^0-9A-Fa-f]', '', raw).upper()
        if len(compact) != 12:
            return ''
        return compact

    def display_mac(value):
        mac = normalize_mac(value)
        if not mac:
            return str(value or '').strip()
        return ':'.join(mac[i:i + 2] for i in range(0, 12, 2))

    def parse_datetime(value):
        text = str(value or '').strip()
        if not text:
            return None
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d'):
            try:
                return datetime.datetime.strptime(text[:19], fmt)
            except Exception:
                pass
        return None

    def read_csv(file_storage):
        data = file_storage.read()
        for encoding in ('utf-8-sig', 'gb18030', 'gbk'):
            try:
                text = data.decode(encoding)
                raw_rows = list(csv.reader(io.StringIO(text)))
                headers, data_rows = detect_table(raw_rows)
                return data_rows, headers
            except Exception:
                continue
        raise ValueError('CSV 编码无法识别，请使用 UTF-8 或 GBK 编码导出')

    def read_xlsx(file_storage):
        file_storage.stream.seek(0)
        workbook = openpyxl.load_workbook(file_storage.stream, read_only=True, data_only=True)
        all_headers = []
        all_data = []
        for sheet in workbook.worksheets:
            raw_rows = list(sheet.iter_rows(values_only=True))
            headers, data = detect_table(raw_rows, sheet.title)
            if headers and not all_headers:
                all_headers = headers
            for header in headers:
                if header not in all_headers:
                    all_headers.append(header)
            all_data.extend(data)
        return all_data, all_headers

    def detect_table(raw_rows, sheet_name=''):
        header_index = -1
        headers = []
        best_score = 0
        for index, values in enumerate(raw_rows[:20]):
            candidate = [str(cell or '').strip() for cell in values]
            normalized = [normalize_header(cell) for cell in candidate]
            score = 0
            for key in ('ip', 'mac'):
                if any(normalize_header(alias) in normalized for alias in FIELD_ALIASES[key]):
                    score += 3
            score += sum(1 for key in ('last_seen', 'switch_name', 'port', 'core_version', 'install_status', 'protection_status') if any(normalize_header(alias) in normalized for alias in FIELD_ALIASES[key]))
            if score > best_score:
                best_score = score
                header_index = index
                headers = candidate
            if score >= 6:
                break
        if header_index < 0 or best_score < 3:
            return [], []
        headers = make_unique_headers(headers)
        data = []
        for values in raw_rows[header_index + 1:]:
            item = {}
            empty = True
            for index, header in enumerate(headers):
                value = values[index] if index < len(values) else ''
                if value not in (None, ''):
                    empty = False
                item[header] = '' if value is None else value
            if not empty:
                item['__sheet_name'] = sheet_name
                data.append(item)
        return headers, data

    def make_unique_headers(headers):
        output = []
        seen = defaultdict(int)
        for index, header in enumerate(headers, start=1):
            name = str(header or '').strip() or f'未命名列{index}'
            seen[name] += 1
            if seen[name] > 1:
                name = f'{name}_{seen[name]}'
            output.append(name)
        return output

    def read_table(file_storage):
        if not file_storage or not file_storage.filename:
            return [], []
        filename = file_storage.filename.lower()
        if filename.endswith('.csv'):
            return read_csv(file_storage)
        if filename.endswith(('.xlsx', '.xlsm')):
            return read_xlsx(file_storage)
        raise ValueError('仅支持 CSV、XLSX、XLSM 文件')

    def template_response(kind):
        columns = TEMPLATE_COLUMNS.get(kind)
        if not columns:
            return jsonify({'status': 'error', 'msg': '未知模板类型'}), 404
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = '准入合规参考模板'
        sheet.append(columns)
        sheet.append(TEMPLATE_SAMPLES[kind])
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill('solid', fgColor='EAF3FF')
        for index, column in enumerate(columns, start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = max(14, len(column) + 8)
        bio = io.BytesIO()
        workbook.save(bio)
        bio.seek(0)
        filename = '9200平台参考模板.xlsx' if kind == 'agent' else '人工审批台账参考模板.xlsx'
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def standardize_rows(rows, headers, source):
        mapping = {key: guess_field(headers, key) for key in FIELD_ALIASES}
        output = []

        def row_value(row, key):
            primary = mapping.get(key)
            if primary and row.get(primary) not in (None, ''):
                return row.get(primary)
            for alias in FIELD_ALIASES.get(key, []):
                alias_norm = normalize_header(alias)
                for header, value in row.items():
                    if normalize_header(header) == alias_norm and value not in (None, ''):
                        return value
            return ''

        for index, row in enumerate(rows, start=2):
            raw_ip = row_value(row, 'ip')
            raw_mac = row_value(row, 'mac')
            item = {
                'source': source,
                'row': index,
                'ip': normalize_ip(raw_ip),
                'raw_ip': str(raw_ip or '').strip(),
                'mac': normalize_mac(raw_mac),
                'raw_mac': str(raw_mac or '').strip(),
                'last_seen': str(row_value(row, 'last_seen')).strip(),
                'switch_name': str(row_value(row, 'switch_name')).strip(),
                'switch_ip': str(row_value(row, 'switch_ip')).strip(),
                'port': str(row_value(row, 'port')).strip(),
                'vlan': str(row_value(row, 'vlan')).strip(),
                'status': str(row_value(row, 'status')).strip(),
                'group': str(row_value(row, 'group')).strip(),
                'owner': str(row_value(row, 'owner')).strip(),
                'department': str(row_value(row, 'department')).strip() or str(row.get('__sheet_name', '')).strip(),
                'device_type': str(row_value(row, 'device_type')).strip(),
                'install_location': str(row_value(row, 'install_location')).strip(),
                'device_purpose': str(row_value(row, 'device_purpose')).strip(),
                'asset_status': str(row_value(row, 'asset_status')).strip(),
                'note': str(row_value(row, 'note')).strip(),
                'core_version': str(row_value(row, 'core_version')).strip(),
                'install_status': str(row_value(row, 'install_status')).strip(),
                'protection_status': str(row_value(row, 'protection_status')).strip(),
                'hostname': str(row_value(row, 'hostname')).strip(),
            }
            item['last_seen_dt'] = parse_datetime(item['last_seen'])
            item['agent_registered'] = '9200' in item['core_version']
            item['agent_trusted'] = item['protection_status'] == '保护'
            item['agent_legal'] = item['agent_registered'] or item['agent_trusted']
            output.append(item)
        return output, mapping

    DEFAULT_EVIDENCE = {
        '无有效IP': 'IP地址管理平台：IP字段为空、非法或属于无效地址段。',
        '同IP多MAC': 'IP地址管理平台：按IP聚合发现多个不同MAC。',
        '同MAC多IP': 'IP地址管理平台：按MAC聚合发现多个不同IP。',
        '私自分配地址': 'IP地址管理平台有记录，人工台账未找到该IP。',
        'IP-MAC不符': 'IP地址管理平台与人工台账按IP比对，MAC不一致。',
        '疑似私改IP': 'IP地址管理平台与人工台账按MAC比对，IP不一致。',
        '未报备入网': 'IP地址管理平台有记录，人工台账无对应IP-MAC审批记录。',
        '9200未合规': 'IP地址管理平台匹配到9200平台记录，但未满足保护或9200版本条件。',
        '未纳入9200': 'IP地址管理平台有记录，9200平台无合法终端记录。',
        '伪装合规': 'IP地址管理平台和9200平台有合法记录，但人工台账无对应审批。',
        '疑似废弃下线': 'IP地址管理平台与9200平台均为旧记录，人工台账无对应审批记录，且同MAC无近期活跃证据。',
    }

    def risk(level, risk_type, row, message, suggestion, evidence=''):
        mac_text = display_mac(row.get('mac') or row.get('raw_mac'))
        return {
            'level': level,
            'type': risk_type,
            'ip': row.get('ip') or row.get('raw_ip') or '-',
            'mac': mac_text,
            'mac_vendor': lookup_mac_vendor(mac_text),
            'switch_name': row.get('switch_name') or '-',
            'switch_ip': row.get('switch_ip') or '-',
            'port': row.get('port') or '-',
            'vlan': row.get('vlan') or '-',
            'last_seen': row.get('last_seen') or '-',
            'status': row.get('status') or '-',
            'owner': row.get('owner') or '-',
            'department': row.get('department') or '-',
            'install_location': row.get('install_location') or '-',
            'device_type': row.get('device_type') or '-',
            'device_purpose': row.get('device_purpose') or '-',
            'asset_status': row.get('asset_status') or '-',
            'note': row.get('note') or '-',
            'evidence': evidence or DEFAULT_EVIDENCE.get(risk_type, '-'),
            'message': message,
            'suggestion': suggestion,
        }

    def latest_key(row):
        return row.get('last_seen_dt') or datetime.datetime.min

    def compact_asset_row(row):
        mac_text = display_mac(row.get('mac') or row.get('raw_mac'))
        return {
            'ip': row.get('ip') or row.get('raw_ip') or '-',
            'mac': mac_text,
            'mac_vendor': lookup_mac_vendor(mac_text),
            'switch_name': row.get('switch_name') or '-',
            'switch_ip': row.get('switch_ip') or '-',
            'port': row.get('port') or '-',
            'vlan': row.get('vlan') or '-',
            'last_seen': row.get('last_seen') or '-',
            'owner': row.get('owner') or '-',
            'department': row.get('department') or '-',
            'install_location': row.get('install_location') or '-',
            'device_type': row.get('device_type') or '-',
            'device_purpose': row.get('device_purpose') or '-',
            'asset_status': row.get('asset_status') or '-',
            'note': row.get('note') or '-',
        }

    def first_lookup_row(rows):
        return rows[0] if rows else None

    def find_agent_row(row, agent_by_ip, agent_by_mac):
        ip = row.get('ip')
        mac = row.get('mac')
        return first_lookup_row(agent_by_ip.get(ip)) or first_lookup_row(agent_by_mac.get(mac) if mac else [])

    def is_older_than(row, cutoff):
        return bool(cutoff and row and row.get('last_seen_dt') and row['last_seen_dt'] < cutoff)

    def is_recent_than(row, cutoff):
        return bool(cutoff and row and row.get('last_seen_dt') and row['last_seen_dt'] >= cutoff)

    def has_recent_mac_evidence(mac, cutoff, *row_groups):
        if not mac or not cutoff:
            return False
        for rows in row_groups:
            for row in rows or []:
                if is_recent_than(row, cutoff):
                    return True
        return False

    def merge_locator(base, *sources):
        merged = dict(base or {})
        for source in sources:
            if not source:
                continue
            for key in (
                'owner', 'department', 'install_location', 'device_type',
                'device_purpose', 'asset_status', 'note', 'hostname'
            ):
                if not merged.get(key) and source.get(key):
                    merged[key] = source.get(key)
        return merged

    def build_agent_lookup(agent_rows):
        legal_ip = set()
        legal_mac = set()
        rows_by_ip = defaultdict(list)
        rows_by_mac = defaultdict(list)
        for row in agent_rows:
            if row.get('ip'):
                rows_by_ip[row['ip']].append(row)
            if row.get('mac'):
                rows_by_mac[row['mac']].append(row)
            if row.get('agent_legal'):
                if row.get('ip'):
                    legal_ip.add(row['ip'])
                if row.get('mac'):
                    legal_mac.add(row['mac'])
        return legal_ip, legal_mac, rows_by_ip, rows_by_mac

    def analyze(ipam_rows, agent_rows, registry_rows, options):
        now = datetime.datetime.now()
        days = int(options.get('days') or 0)
        stale_days = int(options.get('stale_days') or 90)
        online_only = str(options.get('online_only') or '0') == '1'
        cutoff = now - datetime.timedelta(days=days) if days > 0 else None
        stale_cutoff = now - datetime.timedelta(days=stale_days) if stale_days > 0 else None

        actual_rows = []
        skipped_old = 0
        skipped_offline = 0
        for row in ipam_rows:
            if online_only and row.get('status') and row.get('status') != '在线':
                skipped_offline += 1
                continue
            if cutoff and row.get('last_seen_dt') and row['last_seen_dt'] < cutoff:
                skipped_old += 1
                continue
            actual_rows.append(row)

        agent_legal_ip, agent_legal_mac, agent_by_ip, agent_by_mac = build_agent_lookup(agent_rows)
        registry_ip = {row['ip'] for row in registry_rows if row.get('ip')}
        registry_mac = {row['mac'] for row in registry_rows if row.get('mac')}
        registry_pair = {(row['ip'], row['mac']) for row in registry_rows if row.get('ip') and row.get('mac')}
        registry_by_ip = defaultdict(list)
        registry_by_mac = defaultdict(list)
        actual_by_mac = defaultdict(list)
        for row in actual_rows:
            if row.get('mac'):
                actual_by_mac[row['mac']].append(row)
        for row in registry_rows:
            if row.get('ip'):
                registry_by_ip[row['ip']].append(row)
            if row.get('mac'):
                registry_by_mac[row['mac']].append(row)

        risks = []
        ip_to_macs = defaultdict(set)
        mac_to_ips = defaultdict(set)
        latest_by_ip_mac = {}
        for row in actual_rows:
            if not row.get('ip'):
                risks.append(risk('medium', '无有效IP', row, 'IP 地址为空、非法或属于无效地址段', '核查该 MAC 是否真实入网，确认 DHCP 或平台采集状态'))
                continue
            if row.get('ip') and row.get('mac'):
                ip_to_macs[row['ip']].add(row['mac'])
                mac_to_ips[row['mac']].add(row['ip'])
                key = (row['ip'], row['mac'])
                if key not in latest_by_ip_mac or latest_key(row) > latest_key(latest_by_ip_mac[key]):
                    latest_by_ip_mac[key] = row

        for ip, macs in ip_to_macs.items():
            if len(macs) > 1:
                group_rows = [
                    latest_by_ip_mac.get((ip, mac), {'ip': ip, 'mac': mac})
                    for mac in sorted(macs)
                ]
                row = group_rows[0]
                item = risk('high', '同IP多MAC', row, f'同一 IP 对应 {len(macs)} 个 MAC', '排查 IP 冲突、私接设备或 ARP 异常')
                item['group_key'] = ip
                item['group_mode'] = 'ip_to_macs'
                item['group_items'] = [compact_asset_row(group_row) for group_row in group_rows]
                item['group_count'] = len(group_rows)
                risks.append(item)

        for mac, ips in mac_to_ips.items():
            if len(ips) > 1:
                group_rows = [
                    latest_by_ip_mac.get((ip, mac), {'ip': ip, 'mac': mac})
                    for ip in sorted(ips, key=lambda value: tuple(int(part) if part.isdigit() else 0 for part in str(value).split('.')))
                ]
                row = group_rows[0]
                item = risk('high', '同MAC多IP', row, f'同一 MAC 使用 {len(ips)} 个 IP', '排查终端私改 IP、双地址配置或虚拟网卡')
                item['group_key'] = display_mac(mac)
                item['group_mode'] = 'mac_to_ips'
                item['group_items'] = [compact_asset_row(group_row) for group_row in group_rows]
                item['group_count'] = len(group_rows)
                risks.append(item)

        for row in actual_rows:
            ip = row.get('ip')
            mac = row.get('mac')
            if not ip:
                continue
            agent_row = find_agent_row(row, agent_by_ip, agent_by_mac)
            registry_matched = (ip, mac) in registry_pair if mac else ip in registry_ip
            recent_mac_evidence = has_recent_mac_evidence(
                mac,
                stale_cutoff,
                actual_by_mac.get(mac),
                agent_by_mac.get(mac),
            )
            if registry_rows and agent_row and not registry_matched and not recent_mac_evidence and is_older_than(row, stale_cutoff) and is_older_than(agent_row, stale_cutoff):
                risks.append(risk(
                    'medium',
                    '疑似废弃下线',
                    merge_locator(row, agent_row),
                    f'IP管理与9200平台均存在超过 {stale_days} 天的旧记录，且未匹配人工台账',
                    '优先核实是否历史残留；确认已下线后清理平台旧记录，仍在使用则补登记',
                    f'IP地址管理平台最近上线：{row.get("last_seen") or "-"}；9200平台更新时间：{agent_row.get("last_seen") or "-"}；人工台账未匹配IP-MAC；同MAC未发现阈值内近期记录。'
                ))
                continue
            if registry_rows:
                if ip not in registry_ip:
                    risks.append(risk('high', '私自分配地址', row, '该 IP 未出现在人工审批台账中', '核查责任人，确认是否补登记或断网处理'))
                if mac and (ip, mac) not in registry_pair:
                    if ip in registry_ip and mac not in {r.get('mac') for r in registry_by_ip[ip]}:
                        registry_row = first_lookup_row(registry_by_ip[ip])
                        risks.append(risk('high', 'IP-MAC不符', merge_locator(row, registry_row), '该 IP 的实际 MAC 与台账登记不一致', '优先现场核查，防止盗用固定 IP'))
                    elif mac in registry_mac and ip not in {r.get('ip') for r in registry_by_mac[mac]}:
                        registry_row = first_lookup_row(registry_by_mac[mac])
                        risks.append(risk('medium', '疑似私改IP', merge_locator(row, registry_row), '该 MAC 出现在台账中，但当前 IP 与登记不一致', '核查终端是否私自修改地址'))
                    else:
                        risks.append(risk('high', '未报备入网', row, '实际入网终端未匹配到人工审批 IP-MAC 记录', '确认是否已批准入网，未批准则按非法终端处理'))
            if agent_rows and ip not in agent_legal_ip and (not mac or mac not in agent_legal_mac):
                has_agent_record = bool(agent_row)
                if has_agent_record:
                    risks.append(risk('high', '9200未合规', merge_locator(row, agent_row), '该终端存在 9200 平台记录，但未满足“保护状态=保护”或“核心版本包含9200”', '核查终端是否完成客户端版本升级或人工信任审核'))
                else:
                    risks.append(risk('high', '未纳入9200', row, '实际入网终端未匹配到 9200 合法终端记录', '安装保护软件或核查是否允许入网'))
            if agent_rows and registry_rows and (ip in agent_legal_ip or (mac and mac in agent_legal_mac)) and (not mac or (ip, mac) not in registry_pair):
                risks.append(risk('medium', '伪装合规', merge_locator(row, agent_row), '终端存在 9200 记录，但未匹配人工审批台账', '补充审批责任链，确认是否绕过入网登记'))

        registered_inactive = []
        if registry_rows:
            actual_pairs = {(row.get('ip'), row.get('mac')) for row in actual_rows if row.get('ip') and row.get('mac')}
            for row in registry_rows:
                if row.get('ip') and row.get('mac') and (row['ip'], row['mac']) not in actual_pairs:
                    registered_inactive.append(row)

        level_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
        risks.sort(key=lambda item: (level_order.get(item['level'], 9), item['type'], item['ip']))
        risk_by_type = defaultdict(int)
        for item in risks:
            risk_by_type[item['type']] += 1
        summary = {
            'ipam_total': len(ipam_rows),
            'agent_total': len(agent_rows),
            'agent_legal': sum(1 for row in agent_rows if row.get('agent_legal')),
            'agent_registered': sum(1 for row in agent_rows if row.get('agent_registered')),
            'agent_trusted': sum(1 for row in agent_rows if row.get('agent_trusted')),
            'registry_total': len(registry_rows),
            'analyzed_actual': len(actual_rows),
            'skipped_offline': skipped_offline,
            'skipped_old': skipped_old,
            'risk_total': len(risks),
            'high': sum(1 for item in risks if item['level'] == 'high'),
            'medium': sum(1 for item in risks if item['level'] == 'medium'),
            'low': sum(1 for item in risks if item['level'] == 'low'),
            'registered_inactive': len(registered_inactive),
            'risk_by_type': dict(sorted(risk_by_type.items(), key=lambda pair: (-pair[1], pair[0]))),
            'stale_days': stale_days,
        }
        return {
            'summary': summary,
            'risks': risks,
            'risk_types': list(summary['risk_by_type'].keys()),
            'risk_truncated': False,
            'registered_inactive_sample': [compact_asset_row(row) for row in registered_inactive[:100]],
        }

    @bp.route('/api/compliance/analyze', methods=['POST'])
    @login_required
    def api_compliance_analyze():
        try:
            ipam_file = request.files.get('ipam_file')
            if not ipam_file or not ipam_file.filename:
                return jsonify({'status': 'error', 'msg': '请上传 IP 地址管理平台导出文件'}), 400

            datasets = {}
            mappings = {}
            for key, label in [
                ('ipam_file', 'ipam'),
                ('agent_file', 'agent'),
                ('registry_file', 'registry'),
            ]:
                uploaded = request.files.get(key)
                if uploaded and uploaded.filename:
                    rows, headers = read_table(uploaded)
                    datasets[label], mappings[label] = standardize_rows(rows, headers, label)
                else:
                    datasets[label], mappings[label] = [], {}

            result = analyze(
                datasets['ipam'],
                datasets['agent'],
                datasets['registry'],
                {
                    'days': request.form.get('days') or 0,
                    'stale_days': request.form.get('stale_days') or 90,
                    'online_only': request.form.get('online_only') or '0',
                },
            )
            result['mappings'] = mappings
            result['history_id'] = db.save_compliance_analysis_run(
                getattr(current_user, 'id', '') and getattr(current_user, 'username', ''),
                {
                    'ipam': ipam_file.filename,
                    'agent': request.files.get('agent_file').filename if request.files.get('agent_file') else '',
                    'registry': request.files.get('registry_file').filename if request.files.get('registry_file') else '',
                },
                result.get('summary') or {},
                result.get('risks') or [],
            )
            return jsonify({'status': 'success', 'data': result})
        except Exception as e:
            return internal_error('准入合规分析失败，请检查文件格式和字段内容', e)

    @bp.route('/api/compliance/template/<kind>', methods=['GET'])
    @login_required
    def api_compliance_template(kind):
        return template_response(kind)

    @bp.route('/api/compliance/history', methods=['GET'])
    @login_required
    def api_compliance_history():
        try:
            limit = request.args.get('limit') or 20
            return jsonify({'status': 'success', 'data': db.list_compliance_analysis_runs(limit)})
        except Exception as e:
            return internal_error('读取准入合规历史记录失败', e)

    @bp.route('/api/compliance/history/<int:run_id>', methods=['GET'])
    @login_required
    def api_compliance_history_detail(run_id):
        try:
            item = db.get_compliance_analysis_run(run_id)
            if not item:
                return jsonify({'status': 'error', 'msg': '历史记录不存在'}), 404
            return jsonify({
                'status': 'success',
                'data': {
                    'summary': item.get('summary') or {},
                    'risks': item.get('risks') or [],
                    'risk_types': list((item.get('summary') or {}).get('risk_by_type') or {}),
                    'history': {
                        'id': item.get('id'),
                        'created_at': item.get('created_at'),
                        'username': item.get('username'),
                        'ipam_filename': item.get('ipam_filename'),
                        'agent_filename': item.get('agent_filename'),
                        'registry_filename': item.get('registry_filename'),
                    },
                },
            })
        except Exception as e:
            return internal_error('读取准入合规历史详情失败', e)

    return bp
