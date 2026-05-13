import datetime
import io

import openpyxl
from flask import Blueprint, jsonify, request, send_file
from flask_login import current_user, login_required


def create_asset_manage_blueprint(
    db,
    get_json_data,
    require_fields,
    normalize_ip,
    normalize_port,
    normalize_vendor,
    normalize_switch_role,
    json_error,
    internal_error,
    permission_required,
    has_permission,
):
    bp = Blueprint('asset_manage', __name__)

    def parse_switch_import_workbook(file_storage, apply=False):
        wb = openpyxl.load_workbook(file_storage, data_only=True)
        sheet = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        required_cols = ['设备名称', 'IP地址', '端口', '用户名', '密码', '厂商']
        col_indices = {}
        for req in required_cols:
            if req in headers:
                col_indices[req] = headers.index(req)
            else:
                raise ValueError(f"资产表格缺少必填列头：{req}")

        existing_switches = db.get_all_switches()
        existing_ips = {s['ip'] for s in existing_switches}
        seen_ips = set()
        rows = []
        errors = []
        warnings = []
        created = 0
        skipped = 0
        for row_no, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            raw_ip = row[col_indices['IP地址']]
            if not raw_ip:
                continue
            try:
                ip = normalize_ip(raw_ip)
                port = normalize_port(row[col_indices['端口']] or 22)
                name = str(row[col_indices['设备名称']] or f"Switch_{ip}").strip()
                user = str(row[col_indices['用户名']] or '').strip()
                pwd = str(row[col_indices['密码']] or '').strip()
                vendor = normalize_vendor(row[col_indices['厂商']] or 'h3c')
                role_index = headers.index('角色') if '角色' in headers else None
                role = normalize_switch_role(row[role_index] if role_index is not None else 'access')
                if not user:
                    raise ValueError('用户名不能为空')
                if not pwd:
                    raise ValueError('密码不能为空')
                status = 'ready'
                note = '可导入'
                if ip in existing_ips:
                    status = 'skip'
                    note = 'IP 已存在，将跳过'
                    skipped += 1
                elif ip in seen_ips:
                    status = 'skip'
                    note = '表格内重复 IP，将跳过'
                    skipped += 1
                else:
                    seen_ips.add(ip)
                    if apply:
                        db.add_switch(name, ip, port, user, pwd, vendor, role)
                        existing_ips.add(ip)
                    created += 1
                rows.append({
                    'row': row_no,
                    'name': name,
                    'ip': ip,
                    'port': port,
                    'username': user,
                    'vendor': vendor,
                    'role': role,
                    'status': status,
                    'note': note,
                })
            except ValueError as exc:
                errors.append({'row': row_no, 'message': str(exc)})
                rows.append({
                    'row': row_no,
                    'name': str(row[col_indices['设备名称']] or ''),
                    'ip': str(raw_ip or ''),
                    'status': 'error',
                    'note': str(exc),
                })
        return {
            'total': len(rows),
            'created': created,
            'skipped': skipped,
            'errors': errors,
            'warnings': warnings,
            'rows': rows[:100],
        }

    @bp.route('/api/switches/export', methods=['GET'])
    @login_required
    def export_switches():
        try:
            include_password = request.args.get('include_password') == '1'
            if include_password and not has_permission('asset.export_sensitive'):
                return json_error('当前账号缺少权限：asset.export_sensitive', 403)
            switches = db.get_all_switches()
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = 'switch_assets'
            headers = ['设备名称', 'IP地址', '端口', '用户名', '厂商', '角色']
            if include_password:
                headers.insert(4, '密码')
            sheet.append(headers)
            for sw in switches:
                row = [
                    sw.get('name') or '',
                    sw.get('ip') or '',
                    sw.get('port') or 22,
                    sw.get('username') or '',
                    sw.get('vendor') or 'h3c',
                    sw.get('role') or 'access',
                ]
                if include_password:
                    row.insert(4, sw.get('password') or '')
                sheet.append(row)
            for column_cells in sheet.columns:
                column_letter = column_cells[0].column_letter
                max_len = max(len(str(cell.value or '')) for cell in column_cells)
                sheet.column_dimensions[column_letter].width = min(max(max_len + 4, 12), 28)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            export_type = 'full_sensitive' if include_password else 'safe'
            db.log_operation(
                current_user.username,
                request.remote_addr,
                'LOCAL',
                '导出设备资产',
                f"type={export_type}, count={len(switches)}",
                '成功',
            )
            return send_file(
                output,
                as_attachment=True,
                download_name=f"switch_assets_{export_type}_{timestamp}.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        except Exception as e:
            return internal_error('导出设备资产失败，请稍后重试', e)

    @bp.route('/api/switches/batch_import_preview', methods=['POST'])
    @login_required
    @permission_required('asset.manage')
    def batch_import_switches_preview():
        try:
            file = request.files.get('file')
            if not file or not file.filename:
                return json_error('请选择资产 Excel 文件')
            result = parse_switch_import_workbook(file, apply=False)
            return jsonify({'status': 'success', 'data': result})
        except ValueError as e:
            return json_error(str(e))
        except Exception as e:
            return internal_error('预校验资产 Excel 失败，请检查文件内容', e)

    @bp.route('/api/switches/add', methods=['POST'])
    @login_required
    @permission_required('asset.manage')
    def api_add_switch():
        try:
            data = get_json_data()
            require_fields(data, ['name', 'ip', 'port', 'user'])
            data['ip'] = normalize_ip(data['ip'])
            data['port'] = normalize_port(data['port'])
            vendor = normalize_vendor(data.get('vendor'))
            role = normalize_switch_role(data.get('role'))
            if db.get_switch_by_ip(data['ip']):
                return jsonify({'status': 'error', 'msg': f"添加失败：IP 地址 {data['ip']} 已存在，请勿重复录入！"})

            db.add_switch(data['name'], data['ip'], data['port'], data.get('user'), data.get('pass'), vendor, role)
            return jsonify({'status': 'success'})
        except ValueError as e:
            return json_error(str(e))
        except Exception as e:
            return internal_error('添加设备失败，请检查输入或稍后重试', e)

    @bp.route('/api/switches/batch_import', methods=['POST'])
    @login_required
    @permission_required('asset.manage')
    def batch_import_switches():
        try:
            file = request.files.get('file')
            if not file or not file.filename:
                return json_error('请选择资产 Excel 文件')
            result = parse_switch_import_workbook(file, apply=True)
            if result['errors']:
                return json_error(f"导入中止：存在 {len(result['errors'])} 条格式错误，请先预校验并修正")
            msg = f"成功导入 {result['created']} 台设备！"
            if result['skipped'] > 0:
                msg += f"（自动跳过 {result['skipped']} 条重复 IP）"
            return jsonify({'status': 'success', 'msg': msg, 'data': result})
        except ValueError as e:
            return json_error(str(e))
        except Exception as e:
            return internal_error('批量导入失败，请检查 Excel 内容后重试', e)

    @bp.route('/api/switches/delete', methods=['POST'])
    @login_required
    @permission_required('asset.manage')
    def del_switch_api():
        try:
            data = get_json_data()
            require_fields(data, ['id'])
            db.delete_switch(int(data['id']))
            return jsonify({'status': 'success'})
        except ValueError as e:
            return json_error(str(e))
        except Exception as e:
            return internal_error('删除设备失败，请稍后重试', e)

    @bp.route('/api/switches/update', methods=['POST'])
    @login_required
    @permission_required('asset.manage')
    def update_switch_api():
        try:
            data = get_json_data()
            require_fields(data, ['id', 'name', 'ip', 'port', 'user'])
            switch_id = int(data['id'])
            name = str(data['name']).strip()
            if not name:
                return json_error('设备名称不能为空')
            ip = normalize_ip(data['ip'])
            port = normalize_port(data['port'])
            username = str(data.get('user') or '').strip()
            raw_password = str(data.get('pass') or '').strip()
            password = raw_password if raw_password else None
            vendor = normalize_vendor(data.get('vendor'))
            role = normalize_switch_role(data.get('role'))

            before = db.get_switch_by_id(switch_id)
            if not before:
                return json_error('设备不存在或已被删除')

            same_ip_switch = db.get_switch_by_ip(ip)
            if same_ip_switch and int(same_ip_switch['id']) != switch_id:
                return json_error(f"修改失败：IP 地址 {ip} 已被其他设备使用")

            updated = db.update_switch(switch_id, name, ip, port, username, password, vendor, role)
            if not updated:
                return json_error('设备不存在或已被删除')

            db.log_operation(
                current_user.username,
                request.remote_addr,
                ip,
                "修改设备资产",
                f"id={switch_id}, {before['ip']} -> {ip}, name={before['name']} -> {name}, vendor={vendor}, role={role}, password={'updated' if password is not None else 'kept'}",
                "成功",
            )
            return jsonify({'status': 'success'})
        except ValueError as e:
            return json_error(str(e))
        except Exception as e:
            return internal_error('修改设备失败，请检查输入或稍后重试', e)

    @bp.route('/api/switches/update_metadata', methods=['POST'])
    @login_required
    @permission_required('asset.manage')
    def update_switch_metadata_api():
        try:
            data = get_json_data()
            require_fields(data, ['id'])
            switch_id = int(data['id'])
            role = normalize_switch_role(data.get('role')) if 'role' in data else None
            db.update_switch_metadata(switch_id, role)
            db.log_operation(
                current_user.username,
                request.remote_addr,
                str(switch_id),
                "更新设备资产属性",
                f"role={role if role is not None else '-'}",
                "成功",
            )
            return jsonify({'status': 'success'})
        except ValueError as e:
            return json_error(str(e))
        except Exception as e:
            return internal_error('更新设备资产属性失败，请稍后重试', e)

    return bp
