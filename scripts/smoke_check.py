import io
import os
import re
import sys
import time
import zipfile

import openpyxl

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

from app import APP_VERSION, app
import database as db
from background_tasks import get_task, list_tasks, submit_task


def expect(condition, message):
    if not condition:
        raise AssertionError(message)


def expect_contains(text, markers, label):
    missing = [marker for marker in markers if marker not in text]
    expect(not missing, f'{label} missing markers: {missing}')


def main():
    checks = []
    with app.test_client() as client:
        login = client.post(
            '/login',
            data={'username': 'admin', 'password': 'admin888'},
            follow_redirects=False,
        )
        expect(login.status_code in (302, 303), f'login failed: {login.status_code}')
        checks.append('login')

        version = client.get('/api/version')
        expect(version.status_code == 200, f'/api/version {version.status_code}')
        expect(version.get_json().get('status') == 'success', '/api/version json status')
        checks.append('version')

        index = client.get('/')
        text = index.get_data(as_text=True)
        expect(index.status_code == 200, f'index {index.status_code}')
        expect_contains(
            text,
            [
                'dashboard_area',
                f'/static/css/netmaster-tabler.css?v={APP_VERSION}',
                '/static/vendor/tabler/1.4.0/tabler.min.css',
                '/static/vendor/tabler/1.4.0/tabler.min.js',
                'id="dashboard-tab-btn"',
                'id="alarm-tab-btn"',
                'id="task-tab-btn"',
                'id="terminal-list-tab-btn"',
                'id="compliance-tab-btn"',
                'id="manage-tab-btn"',
                'id="users-tab-btn"',
                'id="data-tab-btn"',
                'id="settings-tab-btn"',
                'id="help-tab-btn"',
                'role-permission-card',
                'nm-icon-btn',
                'nm-status-badge',
                'nm-preview-table',
                'Excel 批量部署',
                '配置差异比对',
                '终端漫游',
                '准入合规分析',
            ],
            'admin page',
        )
        expect('cdn' + '.jsdelivr' not in text and 'un' + 'pkg' not in text, 'external CDN marker found in index')
        checks.append('dashboard-page')

        with open(os.path.join(ROOT_DIR, 'static', 'css', 'netmaster-tabler.css'), 'r', encoding='utf-8') as css_file:
            css_text = css_file.read()
        expect('padding-left: 260px' in css_text, 'main content should reserve fixed-sidebar space')
        expect('grid-template-columns: 236px minmax(0, 1fr)' not in css_text, 'fixed sidebar must not use container grid columns')
        expect('grid-row: 1 / span 999' not in css_text, 'fixed sidebar must not reserve implicit grid rows')
        checks.append('layout-css')

        users = client.get('/api/users')
        expect(users.status_code == 200, f'/api/users {users.status_code}')
        expect(users.get_json().get('status') == 'success', '/api/users json status')
        checks.append('users')

        switches = client.get('/api/switches')
        expect(switches.status_code == 200, f'/api/switches {switches.status_code}')
        expect(switches.get_json().get('status') == 'success', '/api/switches json status')
        checks.append('switches')

        template = client.get('/api/switches/import_template')
        expect(template.status_code == 200, f'import template {template.status_code}')
        wb = openpyxl.load_workbook(io.BytesIO(template.data), read_only=True)
        headers = [cell.value for cell in next(wb.active.iter_rows(max_row=1))]
        expect('角色' in headers, 'switch import template missing role column')
        checks.append('switch-template')

        export = client.get('/api/data_export')
        expect(export.status_code == 200, f'/api/data_export {export.status_code}')
        package = zipfile.ZipFile(io.BytesIO(export.data), 'r')
        names = set(package.namelist())
        expect('net_assets.db' in names, 'data package missing net_assets.db')
        expect('manifest.json' in names, 'data package missing manifest.json')
        checks.append('data-export')

        audit_options = client.get('/api/audit_logs/options')
        expect(audit_options.status_code == 200, f'/api/audit_logs/options {audit_options.status_code}')
        expect(audit_options.get_json().get('status') == 'success', '/api/audit_logs/options json status')
        checks.append('audit-options')

        backup_files = client.get('/api/backup_files?limit=5')
        expect(backup_files.status_code == 200, f'/api/backup_files {backup_files.status_code}')
        expect(backup_files.get_json().get('status') == 'success', '/api/backup_files json status')
        checks.append('backup-files')

        help_page = client.get('/api/help')
        expect(help_page.status_code == 200, f'/api/help {help_page.status_code}')
        help_json = help_page.get_json()
        expect(help_json.get('status') == 'success', '/api/help json status')
        expect('端口 IP+MAC 绑定' in help_json.get('data', {}).get('content', ''), '/api/help missing guide content')
        checks.append('help')

        compliance_empty = client.post('/api/compliance/analyze', data={})
        expect(compliance_empty.status_code == 400, f'/api/compliance/analyze empty upload {compliance_empty.status_code}')
        expect(compliance_empty.get_json().get('status') == 'error', '/api/compliance/analyze empty upload json status')
        for kind in ('agent', 'registry'):
            compliance_template = client.get(f'/api/compliance/template/{kind}')
            expect(compliance_template.status_code == 200, f'/api/compliance/template/{kind} {compliance_template.status_code}')
            template_wb = openpyxl.load_workbook(io.BytesIO(compliance_template.data), read_only=True)
            template_headers = [cell.value for cell in next(template_wb.active.iter_rows(max_row=1))]
            if kind == 'registry':
                expect('设备ip' in template_headers and '设备mac' in template_headers, f'compliance {kind} template missing device IP/MAC')
            else:
                expect('IP地址' in template_headers and 'MAC地址' in template_headers, f'compliance {kind} template missing IP/MAC')
        checks.append('compliance')

        task = submit_task(
            'smoke persistent task',
            lambda task_id: {'ok': True, 'task_id': task_id},
            category='smoke',
            actor='smoke',
            target='local',
        )
        task_id = task.get('id')
        task_snapshot = None
        persisted = None
        for _ in range(40):
            task_snapshot = get_task(task_id)
            persisted = db.get_runtime_task(task_id)
            if (
                task_snapshot
                and task_snapshot.get('status') == 'success'
                and persisted
                and persisted.get('status') == 'success'
            ):
                break
            time.sleep(0.05)
        expect(task_snapshot and task_snapshot.get('status') == 'success', 'runtime task did not finish')
        expect(persisted and persisted.get('status') == 'success', 'runtime task was not persisted')
        listed = list_tasks(limit=20, category='smoke')
        expect(any(row.get('id') == task_id for row in listed), 'runtime task missing from list')
        checks.append('runtime-task-persistence')

        boundary_posts = [
            '/test_connection',
            '/get_interfaces',
            '/get_port_info',
            '/api/port_probe/start',
            '/api/port_probe_asset/start',
            '/api/port_snapshots/collect',
            '/set_interface_description',
            '/bind_port',
            '/del_port_binding',
            '/api/terminal_binding_lookup',
            '/api/migrate_terminal',
            '/api/execute_excel_group',
            '/api/execute_excel_row',
            '/get_acl',
            '/add_acl',
            '/del_acl',
            '/save_config',
            '/api/switch_alarm_logs',
            '/api/switch_alarm_logs/start',
            '/api/alarm_state/update',
        ]
        original_log_operation = db.log_operation
        db.log_operation = lambda *args, **kwargs: None
        try:
            for url in boundary_posts:
                response = client.post(url, json={})
                body = response.get_json(silent=True) or {}
                expect(response.status_code in (400, 409), f'{url} empty payload expected 400/409, got {response.status_code}')
                expect(body.get('status') in ('error', 'busy'), f'{url} empty payload json status')
        finally:
            db.log_operation = original_log_operation
        checks.append('write-boundaries')

        active_operator = next(
            (
                user for user in db.list_users()
                if user.get('role') == 'operator' and user.get('status') == 'active'
            ),
            None,
        )
        if active_operator:
            with client.session_transaction() as sess:
                sess['_user_id'] = str(active_operator['id'])
                sess['_fresh'] = True
            operator_page = client.get('/')
            operator_text = operator_page.get_data(as_text=True)
            expect(operator_page.status_code == 200, f'operator index {operator_page.status_code}')
            for tab_id, label in [
                ('users-tab-btn', 'user management'),
                ('data-tab-btn', 'data backup'),
                ('settings-tab-btn', 'settings'),
            ]:
                item_match = re.search(rf'<li class="([^"]*d-none[^"]*)"[^>]*>\s*<button[^>]+id="{tab_id}"', operator_text)
                expect(item_match, f'operator can see {label} menu')
            side_action_match = re.search(r'nav-side-actions|nav-side-action', operator_text)
            expect(not side_action_match, 'side batch backup action should be removed')
            denied = client.get('/api/users')
            expect(denied.status_code == 403, f'operator /api/users should be 403, got {denied.status_code}')
            denied_export = client.get('/api/data_export')
            expect(denied_export.status_code == 403, f'operator /api/data_export should be 403, got {denied_export.status_code}')
            denied_sensitive_export = client.get('/api/switches/export?include_password=1')
            expect(
                denied_sensitive_export.status_code == 403,
                f'operator sensitive switch export should be 403, got {denied_sensitive_export.status_code}',
            )
            safe_export = client.get('/api/switches/export')
            expect(safe_export.status_code == 200, f'operator safe switch export should be 200, got {safe_export.status_code}')
            denied_settings = client.post('/api/settings/update', json={})
            expect(denied_settings.status_code == 403, f'operator /api/settings/update should be 403, got {denied_settings.status_code}')
            allowed_tasks = client.get('/api/runtime_tasks?limit=1')
            expect(allowed_tasks.status_code == 200, f'operator /api/runtime_tasks should be 200, got {allowed_tasks.status_code}')
            checks.append('operator-permissions')
        else:
            checks.append('operator-permissions-skipped')

    print('Smoke checks passed:', ', '.join(checks))
    return 0


if __name__ == '__main__':
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f'Smoke checks failed: {exc}', file=sys.stderr)
        raise
